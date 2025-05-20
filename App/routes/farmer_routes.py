import requests, random, string, os, secrets, traceback
import pandas as pd

from flask import Blueprint, flash, render_template, json, redirect, url_for, current_app, request, jsonify, send_file
from flask_login import login_required, current_user
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from datetime import datetime, date, timedelta
from babel.numbers import format_currency
from babel.dates import format_date
from werkzeug.utils import secure_filename
from typing import Optional, Tuple
from openpyxl import load_workbook, Workbook
from functools import wraps
from io import BytesIO
from decimal import Decimal

from App import db, cache
from App.models import Kebun, DataPangan, KebunKomoditas, Komoditas, Order, OrderItem, Product, PetaniProfile, Transaction, User, ProductBatch
from App.services.xendit_client import XenditClient, XenditError

farmer = Blueprint('farmer', __name__)

# Constants for API URL parameters
KAB_KOTA = 458  # Ternate
KOMODITAS_ID = 3
TARGET_KOMODITAS = ["Cabai Merah Keriting", "Cabai Rawit Merah", "Bawang Merah"]
REPORT_ALLOWED_EXTENSIONS = {'xlsx'}
PICTURE_ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}

# Constants for fetching price data from API
API_URL = "https://www.bi.go.id/hargapangan/WebSite/Home/GetGridData1"
COMMODITY_IDS = {
    "Cabai Rawit Merah": "8_16",
    "Cabai Merah Keriting": "7_14",
    "Bawang Merah": "5"
}
PRICE_TYPE_ID = 1
JENIS_ID = 1
PERIOD_ID = 1
PROVINCE_ID = 32

# IMPORT FUNCTION
# Constants
REPORT_ALLOWED_EXTENSIONS = {'xlsx'}
MAX_CONTENT_LENGTH = 5 * 1024 * 1024  # 5MB limit

def role_required(role):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated or not current_user.has_role('petani') or not current_user.is_confirmed:
                flash(f'Hanya {role} yang dapat memiliki izin untuk mengakses halaman ini!', 'warning')
                return redirect(url_for('public.index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def picture_allowed_file(filename):
    return '.' in filename and \
            filename.rsplit('.', 1)[1].lower() in PICTURE_ALLOWED_EXTENSIONS

def report_allowed_file(filename):
    return '.' in filename and \
            filename.rsplit('.', 1)[1].lower() in REPORT_ALLOWED_EXTENSIONS

def generate_unique_id(prefix="KR_", string_length=2, number_length=4):
    """
    Generates a unique ID in the format KR_AB1234.

    Args:
        prefix: The static identifier prefix (default: "KR_").
        string_length: The length of the random string part (default: 2).
        number_length: The length of the random number part (default: 4).

    Returns:
        A unique ID string.
    """
    random_string = ''.join(random.choices(string.ascii_uppercase, k=string_length))
    random_number = ''.join(random.choices(string.digits, k=number_length))
    unique_id = f"{prefix}{random_string}{random_number}"
    return unique_id

# Helper function to fetch commodity data from API
def fetch_commodity_data(target_date, commodity_id):
    params = {
        "tanggal": target_date,
        "commodity": commodity_id,
        "priceType": PRICE_TYPE_ID,
        "isPasokan": 1,
        "jenis": JENIS_ID,
        "periode": PERIOD_ID,
        "provId": PROVINCE_ID,
        "_": int(datetime.now().timestamp() * 1000)
    }

    try:
        response = requests.get(API_URL, params=params)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"Error fetching data: {e}")
        return None

# Helper function to fetch and format price data from API
def fetch_price_data(start_date, end_date):
    """Fetches price data from the API and formats it for display.

    Args:
        start_date (str): The starting date in YYYY-MM-DD format.
        end_date (str): The ending date in YYYY-MM-DD format.

    Returns:
        list: A list of dictionaries containing formatted price data.
    """

    url = f"https://panelharga.badanpangan.go.id/data/kabkota-range-by-levelharga/{KAB_KOTA}/{KOMODITAS_ID}/{start_date}/{end_date}"

    try:
        response = requests.get(url)
        response.raise_for_status()  # Raise an exception for HTTP errors

        data = response.json()

        table_data = []
        for item in data["data"]:
            if item["name"] in TARGET_KOMODITAS:
                for date_data in item["by_date"]:
                    date_obj = datetime.strptime(date_data["date"], "%Y-%m-%d")
                    formatted_date = date_obj.strftime("%d/%m/%Y")
                    geomean_value = date_data["geomean"]

                    # Menyederhanakan format harga
                    formatted_price = "-" if geomean_value == "-" else format_currency(float(geomean_value), "IDR", locale="id_ID", decimal_quantization=False)[:-3]

                    table_data.append({
                        "date": formatted_date,
                        "name": item["name"],
                        "price": formatted_price
                    })

        return table_data

    except requests.exceptions.RequestException as e:
        flash(f"Error fetching data: {e}", category='error')
        return []  # Return an empty list on error

@farmer.route('/api/productivity/cabai-rawit', methods=['GET'])
def get_cabai_productivity():
    try:
        # Get latest harvests for Cabai Rawit Merah
        cabai_harvests = (db.session.query(DataPangan)
            .join(Komoditas)
            .filter(
                Komoditas.nama == 'Cabai Rawit',
                DataPangan.tanggal_panen.isnot(None),
                DataPangan.jml_panen.isnot(None)
            )
            .order_by(DataPangan.tanggal_panen.desc())
            .limit(2)
            .all())

        if len(cabai_harvests) < 2:
            return jsonify({'productivity': 0, 'trend': 'none'})

        # Latest harvest is first due to DESC order
        current_harvest = cabai_harvests[0].jml_panen
        previous_harvest = cabai_harvests[1].jml_panen

        if previous_harvest <= 0:
            return jsonify({'productivity': 0, 'trend': 'none'})

        productivity = ((current_harvest - previous_harvest) / previous_harvest * 100)

        return jsonify({
            'productivity': round(productivity, 1),
            'trend': 'up' if productivity > 0 else 'down' if productivity < 0 else 'none'
        })
    except Exception as e:
        current_app.logger.error(f"Error calculating productivity: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

@farmer.route('/api/calendar-data', methods=['GET'])
def get_calendar_data():
    try:
        plantings = (db.session.query(DataPangan)
            .join(Kebun)
            .join(Komoditas)
            .filter(
                DataPangan.tanggal_bibit.isnot(None),
                DataPangan.is_deleted == False,
                DataPangan.status == 'Penanaman',  # Only show plantings that haven't been harvested
                DataPangan.user_id == current_user.id
            )
            .all())

        calendar_data = []
        today = datetime.now().date()

        for planting in plantings:
            estimated_harvest = planting.tanggal_bibit + timedelta(days=60)
            days_remaining = (estimated_harvest - today).days

            if days_remaining > 0:  # Only include future harvests
                calendar_data.append({
                    'date': estimated_harvest.isoformat(),
                    'kebun': planting.kebun.nama,
                    'komoditas': planting.komoditas_info.nama,
                    'days_remaining': days_remaining
                })

        return jsonify(calendar_data)
    except Exception as e:
        current_app.logger.error(f"Error getting calendar data: {str(e)}")
        return jsonify([])

@farmer.route('/api/daily-productivity/cabai-rawit', methods=['GET'])
def get_daily_productivity():
    try:
        # Get all harvests for Cabai Rawit Merah
        harvests = (db.session.query(DataPangan)
            .join(Komoditas)
            .filter(
                Komoditas.nama == 'Cabai Rawit',
                DataPangan.tanggal_panen.isnot(None),
                DataPangan.jml_panen.isnot(None)
            )
            .order_by(DataPangan.tanggal_panen)
            .all())

        productivity_data = []
        for i in range(1, len(harvests)):
            current = harvests[i].jml_panen
            previous = harvests[i-1].jml_panen

            if previous > 0:
                change = ((current - previous) / previous * 100)
                productivity_data.append({
                    'date': harvests[i].tanggal_panen.isoformat(),
                    'value': round(change, 1)
                })

        return jsonify(productivity_data)
    except Exception as e:
        current_app.logger.error(f"Error getting productivity data: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

@farmer.route('/api/get-price-data', methods=['GET'])
def getpricedata():

    target_date = datetime.today().strftime("%d/%m/%Y")
    all_data = []

    for commodity_name, commodity_id in COMMODITY_IDS.items():
        data = fetch_commodity_data(target_date, commodity_id)
        if data and data.get('data'):
            item = data['data'][0]

            # Parse Indonesian month abbreviations
            raw = item["Tanggal"]            # e.g. "09 Mei 25"
            day, mon_id, yy = raw.split()    # ["09", "Mei", "25"]

            # Map Indonesian month abbreviations to numeric months
            MONTHS = {
                'Jan':'01','Feb':'02','Mar':'03','Apr':'04',
                'Mei':'05','Jun':'06','Jul':'07','Agu':'08',
                'Sep':'09','Okt':'10','Nov':'11','Des':'12'
            }

            # Format the date properly
            yyyy = '20' + yy                # "25" → "2025"
            formatted_date = f"{day}/{MONTHS[mon_id]}/{yyyy}"

            all_data.append({
                "date": formatted_date,
                "name": commodity_name,
                "price": format_currency(item["Nilai"], "IDR", locale="id_ID", decimal_quantization=False)[:-3]
            })

    return jsonify(all_data)

@farmer.route('/api/price-data', methods=['GET', 'POST'])
@cache.memoize(timeout=300)
def get_price_data():
    kab_kota = request.args.get('kab_kota')
    komoditas_id = request.args.get('komoditas_id')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    url = f"https://panelharga.badanpangan.go.id/data/kabkota-range-by-levelharga/{KAB_KOTA}/{KOMODITAS_ID}/{start_date}/{end_date}"

    try:
        response = requests.get(url)
        response.raise_for_status()  # Raise an error for bad responses
        return jsonify(response.json()), 200
    except requests.exceptions.RequestException as e:
        return jsonify({"error": str(e)}), 500

@farmer.route('/petani')
@login_required
@role_required('petani')
def index():
    # Query existing
    data_pangan = DataPangan.query.filter_by(
        user_id=current_user.id,
        status='Penanaman',
        is_deleted=False
    ).all()

    # Query untuk menghitung produktivitas
    harvested_data = DataPangan.query.filter(
        DataPangan.user_id == current_user.id,
        DataPangan.status == 'Panen',
        DataPangan.is_deleted == False,
        DataPangan.jml_panen.isnot(None),
        DataPangan.tanggal_panen.isnot(None)
    ).order_by(DataPangan.tanggal_panen.desc()).limit(2).all()

    # Hitung perubahan produktivitas
    productivity_change = 0
    if len(harvested_data) >= 2:
        current_harvest = harvested_data[0].jml_panen
        previous_harvest = harvested_data[1].jml_panen
        if previous_harvest > 0:  # Hindari pembagian dengan nol
            productivity_change = ((current_harvest - previous_harvest) / previous_harvest) * 100

    # Total panen (dari kode yang sudah ada)
    all_data = DataPangan.query.filter_by(
        user_id=current_user.id,
        is_deleted=False
    ).all()
    total_panen = sum(prod.jml_panen for prod in all_data if prod.jml_panen)

    # Kode untuk harvest_data tetap sama
    today = date.today()
    harvest_data = []
    next_harvest_days = None

    for estPanen in data_pangan:
        if estPanen.estimasi_panen:
            est_date = datetime.strptime(estPanen.estimasi_panen.strftime('%Y-%m-%d'), '%Y-%m-%d').date()
            days_remaining = (est_date - today).days
            if days_remaining > 0:
                harvest_data.append({
                    'date': estPanen.estimasi_panen,
                    'days_remaining': days_remaining,
                    'kebun': estPanen.kebun.nama,
                    'komoditas': estPanen.komoditas_info.nama
                })
                if next_harvest_days is None or days_remaining < next_harvest_days:
                    next_harvest_days = days_remaining

    return render_template('farmer/index.html',
                            total_panen=total_panen,
                            harvest_data=json.dumps(harvest_data),
                            next_harvest_days=next_harvest_days,
                            productivity_change=round(productivity_change, 1),
                            page='farmer_index')

@farmer.route('/petani/manajemen-produksi', methods=['GET', 'POST'])
@login_required
@role_required('petani')
def manajemen_produksi():
    kebun = Kebun.query.filter_by(user_id=current_user.id, is_deleted=False).all()
    komoditas = Komoditas.query.filter_by(is_deleted=False).all()

    # Ambil parameter filter dari request
    field_id = request.args.get('field')
    commodity_id = request.args.get('commodity')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    # Pagination parameters
    page = request.args.get('page', 1, type=int)  # Default to page 1
    per_page = request.args.get('per_page', 10, type=int)  # Default to 10 items per page

    # Ambil data produksi 3 bulan terakhir, termasuk yang masih dalam status "Penanaman"
    three_months_ago = datetime.now() - timedelta(days=90)
    query = DataPangan.query.filter(
        DataPangan.user_id == current_user.id,
        DataPangan.is_deleted == False,
        DataPangan.tanggal_bibit >= three_months_ago
    )

    # Terapkan filter jika ada
    if field_id:
        query = query.filter(DataPangan.kebun_id == field_id)
    if commodity_id:
        query = query.filter(DataPangan.komoditas_id == commodity_id)
    if start_date:
        query = query.filter(DataPangan.tanggal_bibit >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(DataPangan.tanggal_bibit <= datetime.strptime(end_date, '%Y-%m-%d'))

    # Pagination
    produksi = query.paginate(page=page, per_page=per_page)

    # Hitung total akumulasi produksi
    total_produksi = sum(prod.jml_panen for prod in produksi.items if prod.jml_panen is not None)

    # Format data untuk chart
    chart_data = []
    for prod in produksi.items:
        chart_data.append({
            'kebun': prod.kebun.nama,
            'komoditas': prod.komoditas_info.nama,
            'tanggal_panen': format_date(prod.tanggal_panen, format='d MMM Y', locale='id') if prod.tanggal_panen else None,
            'hasil_panen': prod.jml_panen
        })

    # Hitung persentase perubahan produktivitas
    productivity_changes = {}
    for kom in komoditas:
        panen_data = sorted(
            [p for p in produksi.items if p.komoditas_id == kom.id and p.tanggal_panen is not None],
            key=lambda x: x.tanggal_panen
        )
        if len(panen_data) >= 2:
            current = panen_data[-1].jml_panen
            previous = panen_data[-2].jml_panen
            change = ((current - previous) / previous * 100) if previous > 0 else 0
            productivity_changes[kom.id] = {
                'persentase': round(change, 2),
                'trend': 'up' if change >= 0 else 'down'
            }

    user_data = {
        'nama_lengkap': current_user.nama_lengkap,
        'kelurahan': current_user.kelurahan,
        'kota': current_user.kota,
        'luas_lahan': sum(kebun.luas_kebun for kebun in current_user.kebun)
    }

    print(current_user.kebun)

    return render_template('farmer/manajemen_produksi.html',
                            kebun=kebun,
                            komoditas=komoditas,
                            produksi=produksi,
                            total_produksi=total_produksi,
                            chart_data=json.dumps(chart_data),
                            productivity_changes=productivity_changes,
                            format_date=format_date,
                            user_data=user_data,
                            page='farmer_manajemen_produksi'
                            )

@farmer.route('/petani/market/check-new-orders', methods=['GET'])
@login_required
@role_required('petani')
def check_new_orders():
    """Check for new orders since a given timestamp."""
    since = request.args.get('since')

    if not since:
        return jsonify({'success': False, 'message': 'Missing since parameter'}), 400

    try:
        # Parse the timestamp
        since_dt = datetime.fromisoformat(since.replace('Z', '+00:00'))

        # Query new orders
        new_orders = (Order.query
            .join(OrderItem)
            .join(Product)
            .filter(
                Product.seller_id == current_user.id,
                Order.created_at > since_dt
            )
            .order_by(Order.created_at.desc())
            .all())

        # Count pending and paid orders
        pending_count = sum(1 for o in new_orders if o.status == 'pending')
        paid_count = sum(1 for o in new_orders if o.status == 'paid')

        # Format order data for the response
        orders_data = []
        for order in new_orders:
            orders_data.append({
                'id': order.id,
                'created_at': order.created_at.isoformat(),
                'buyer_name': order.buyer.nama_lengkap or order.buyer.username,
                'items_count': len(order.items),
                'total_amount': int(order.total_amount),
                'status': order.status
            })

        # Check for notifications from webhook
        notifications = []
        notification_file = os.path.join(current_app.root_path, 'static', 'notifications', f'seller_{current_user.id}_notifications.json')
        if os.path.exists(notification_file):
            try:
                with open(notification_file, 'r') as f:
                    notifications = json.load(f)

                # Filter notifications by timestamp
                new_notifications = [n for n in notifications if datetime.fromisoformat(n['created_at'].replace('Z', '+00:00')) > since_dt]

                # If we have new notifications but no new orders from the database query,
                # we should check if there are any orders that were updated via webhook
                if new_notifications and not new_orders:
                    for notification in new_notifications:
                        if notification.get('order_id'):
                            # Check if this order belongs to the current user
                            order = Order.query.get(notification['order_id'])
                            if order:
                                # Check if any of the order items are for products sold by this user
                                for item in order.items:
                                    if item.product and item.product.seller_id == current_user.id:
                                        # Add this order to the response
                                        orders_data.append({
                                            'id': order.id,
                                            'created_at': order.created_at.isoformat(),
                                            'buyer_name': order.buyer.nama_lengkap or order.buyer.username,
                                            'items_count': len(order.items),
                                            'total_amount': int(order.total_amount),
                                            'status': order.status
                                        })

                                        # Update counts
                                        if order.status == 'pending':
                                            pending_count += 1
                                        elif order.status == 'paid':
                                            paid_count += 1

                                        break
            except Exception as e:
                current_app.logger.error(f"Error reading notifications: {str(e)}")

        # Get total counts for all orders
        all_pending_count = Order.query.join(OrderItem).join(Product).filter(
            Product.seller_id == current_user.id,
            Order.status == 'pending'
        ).count()

        all_paid_count = Order.query.join(OrderItem).join(Product).filter(
            Product.seller_id == current_user.id,
            Order.status == 'paid'
        ).count()

        return jsonify({
            'success': True,
            'new_orders_count': len(orders_data),
            'pending_count': all_pending_count,
            'paid_count': all_paid_count,
            'orders': orders_data
        })

    except Exception as e:
        current_app.logger.error(f"Error checking for new orders: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@farmer.route('/petani/market/notifications', methods=['GET'])
@login_required
@role_required('petani')
def get_notifications():
    """Get notifications for the current user."""
    try:
        # Check for notifications from webhook
        notifications = []
        notification_file = os.path.join(current_app.root_path, 'static', 'notifications', f'seller_{current_user.id}_notifications.json')
        if os.path.exists(notification_file):
            try:
                with open(notification_file, 'r') as f:
                    notifications = json.load(f)

                # Sort notifications by timestamp (newest first)
                notifications.sort(key=lambda n: n['created_at'], reverse=True)

                # Limit to 10 most recent notifications
                notifications = notifications[:10]

            except Exception as e:
                current_app.logger.error(f"Error reading notifications: {str(e)}")

        return jsonify({
            'success': True,
            'notifications': notifications
        })

    except Exception as e:
        current_app.logger.error(f"Error getting notifications: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@farmer.route('/petani/market', methods=['GET'])
@login_required
@role_required('petani')
def market_manage():
    """List semua produk milik petani + ringkasan pesanan masuk dengan pagination."""
    # Get pagination parameters
    product_page = request.args.get('product_page', 1, type=int)
    order_page = request.args.get('order_page', 1, type=int)
    per_page = 8  # 8 items per page

    # Get search parameters
    product_search = request.args.get('product_search', '')
    order_search = request.args.get('order_search', '')
    order_status_filter = request.args.get('order_status', '')

    # Query products with pagination and search
    product_query = Product.query.filter_by(seller_id=current_user.id)
    if product_search:
        product_query = product_query.filter(Product.nama.ilike(f'%{product_search}%'))

    products_pagination = product_query.order_by(Product.updated_at.desc()).paginate(
        page=product_page, per_page=per_page, error_out=False
    )

    # Query orders with pagination, search, and filter
    order_query = (Order.query
        .join(OrderItem)
        .join(Product)
        .filter(Product.seller_id == current_user.id)
    )

    if order_search:
        # Search by order ID or buyer username
        order_query = order_query.join(User, Order.buyer_id == User.id).filter(
            db.or_(
                Order.id.cast(db.String).ilike(f'%{order_search}%'),
                User.username.ilike(f'%{order_search}%')
            )
        )

    if order_status_filter:
        order_query = order_query.filter(Order.status == order_status_filter)

    orders_pagination = order_query.order_by(Order.created_at.desc()).paginate(
        page=order_page, per_page=per_page, error_out=False
    )

    # Get all orders for statistics (without pagination)
    all_orders = (Order.query
        .join(OrderItem)
        .join(Product)
        .filter(Product.seller_id == current_user.id)
        .all()
    )

    # Calculate statistics
    total_income = sum(
        float(order.total_amount)
        for order in all_orders
        if order.status in ['paid', 'processed', 'shipped', 'completed']
    )

    pending_orders_count = sum(1 for order in all_orders if order.status == 'pending')
    paid_orders_count = sum(1 for order in all_orders if order.status == 'paid')
    processing_orders_count = sum(1 for order in all_orders if order.status in ['processed', 'shipped'])
    completed_orders_count = sum(1 for order in all_orders if order.status == 'completed')

    # Get statistics for products
    active_products_count = Product.query.filter_by(seller_id=current_user.id, is_active=True).count()
    low_stock_products_count = Product.query.filter(
        Product.seller_id == current_user.id,
        Product.is_active == True,
        Product.stok <= 5,
        Product.stok > 0
    ).count()

    # Tambahkan total_income ke context
    return render_template(
        'farmer/market_manage.html',
        products_pagination=products_pagination,
        orders_pagination=orders_pagination,
        total_income=total_income,
        pending_orders_count=pending_orders_count,
        paid_orders_count=paid_orders_count,
        processing_orders_count=processing_orders_count,
        completed_orders_count=completed_orders_count,
        active_products_count=active_products_count,
        low_stock_products_count=low_stock_products_count,
        product_search=product_search,
        order_search=order_search,
        order_status_filter=order_status_filter,
        page='farmer_market'
    )

@farmer.route('/petani/market/transactions', methods=['GET'])
@login_required
@role_required('petani')
def market_transactions():
    """View transaction history"""
    # Get transactions where the farmer is the recipient
    transactions = Transaction.query.filter_by(
        to_user_id=current_user.id
    ).order_by(Transaction.created_at.desc()).all()

    return render_template(
        'farmer/transaction_history.html',
        transactions=transactions,
        page='farmer_market'
    )

@farmer.route('/petani/market/pengaturan', methods=['POST', 'GET'])
@login_required
@role_required('petani')
def market_settings():
    """Update market settings including QRIS and VA settings"""
    upload_folder = current_app.config['UPLOAD_FOLDER']

    if request.method == 'POST':
        try:
            # Get or create petani profile
            profile = current_user.petani_profile
            if not profile:
                # Create profile if it doesn't exist
                profile = PetaniProfile(
                    user_id=current_user.id,
                    unique_id=f"PTN_{secrets.token_hex(4).upper()}"
                )
                db.session.add(profile)
                db.session.flush()

            # Process QRIS settings
            qris_type = request.form.get('qris_type')

            if qris_type == 'static':
                # Process static QRIS picture upload
                if 'qris_picture' in request.files:
                    file = request.files['qris_picture']
                    if file and file.filename:
                        if picture_allowed_file(file.filename):
                            # Generate unique filename
                            filename = secrets.token_hex(8) + '_' + secure_filename(file.filename)

                            # Ensure directory exists
                            qris_dir = os.path.join(upload_folder, 'qris')
                            os.makedirs(qris_dir, exist_ok=True)

                            # Save file
                            file_path = os.path.join(qris_dir, filename)
                            file.save(file_path)

                            # Delete old QRIS file if exists
                            if profile.qris_static:
                                old_file_path = os.path.join(qris_dir, profile.qris_static)
                                if os.path.exists(old_file_path):
                                    os.remove(old_file_path)

                            # Update profile with new QRIS
                            profile.qris_static = filename

                            # Disable dynamic QR
                            profile.qris_dynamic_enabled = False
                            profile.qris_dynamic_id = None
                            profile.qris_dynamic_string = None

                            # Update all products to use the new QRIS
                            products = Product.query.filter_by(seller_id=current_user.id).all()
                            for product in products:
                                product.qris_static = filename

                            flash('QRIS statis berhasil diperbarui!', 'success')
                        else:
                            flash('File yang diizinkan hanya JPG, JPEG, dan PNG.', 'warning')

            elif qris_type == 'dynamic':
                # Enable dynamic QR code generation
                profile.qris_dynamic_enabled = True

                # Create a dynamic QR code using Xendit
                try:
                    # Create a dynamic QR code for the seller
                    qr_data = XenditClient.create_qr_code(
                        seller_id=current_user.id,
                        external_id=f"seller-{current_user.id}-{int(datetime.now().timestamp())}",
                        amount=None  # No amount for seller profile QR
                    )

                    # Store the QR code data in the profile
                    profile.qris_dynamic_id = qr_data.get('id')
                    profile.qris_dynamic_string = qr_data.get('qr_string')

                    # Clear static QR
                    if profile.qris_static:
                        qris_dir = os.path.join(upload_folder, 'qris')
                        old_file_path = os.path.join(qris_dir, profile.qris_static)
                        if os.path.exists(old_file_path):
                            os.remove(old_file_path)
                        profile.qris_static = None

                    flash('QRIS dinamis berhasil diaktifkan!', 'success')
                except XenditError as e:
                    current_app.logger.error(f"Xendit error creating dynamic QR: {str(e)}")
                    flash(f'Gagal membuat QRIS dinamis: {str(e)}', 'danger')
                    return redirect(url_for('farmer.market_manage'))

            # Process VA settings
            va_enabled = request.form.get('va_enabled') == 'on'
            va_bank_code = request.form.get('va_bank_code')

            # Update VA settings
            profile.va_enabled = va_enabled
            if va_bank_code:
                profile.va_bank_code = va_bank_code

            if va_enabled:
                flash('Virtual Account berhasil diaktifkan!', 'success')
            else:
                flash('Virtual Account dinonaktifkan', 'info')

            db.session.commit()

        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error updating market settings: {str(e)}")
            flash(f'Gagal memperbarui pengaturan: {str(e)}', 'danger')

    return redirect(url_for('farmer.market_manage'))

def save_qris_photo(photo_file):
    if not picture_allowed_file(photo_file.filename):
        raise ValueError("Format file tidak diizinkan")
    filename = secure_filename(secrets.token_hex(8) + "_" + photo_file.filename)
    path = os.path.join(current_app.root_path, 'static/uploads/qris', filename)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    photo_file.save(path)
    return url_for('static', filename=f'uploads/qris/{filename}', _external=True)

def save_product_photo(photo_file):
    if not picture_allowed_file(photo_file.filename):
        raise ValueError("Format file tidak diizinkan")
    filename = secure_filename(secrets.token_hex(8) + "_" + photo_file.filename)
    path = os.path.join(filename)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    photo_file.save(path)
    return url_for('static', filename=f'uploads/products/{filename}', _external=True)

@farmer.route('/petani/produksi/check-upcoming-harvests')
@login_required
@role_required('petani')
def check_upcoming_harvests():
    """Check for upcoming harvests and send notifications"""
    today = datetime.now().date()
    upcoming_harvests = DataPangan.query.filter(
        DataPangan.user_id == current_user.id,
        DataPangan.status == 'Penanaman',
        DataPangan.is_deleted == False,
        DataPangan.estimasi_panen.between(today, today + timedelta(days=7))
    ).all()

    notifications = []
    for harvest in upcoming_harvests:
        days_until_harvest = (harvest.estimasi_panen - today).days
        notification = {
            'komoditas': harvest.komoditas_info.nama,
            'kebun': harvest.kebun.nama,
            'estimasi_panen': harvest.estimasi_panen.strftime('%d/%m/%Y'),
            'days_until_harvest': days_until_harvest,
            'produksi_id': harvest.id
        }
        notifications.append(notification)

    return render_template('farmer/upcoming_harvests.html', notifications=notifications)

@farmer.route('/petani/market/product/add', methods=['GET','POST'])
@login_required
@role_required('petani')
def market_add_product():
    """Add a new product to the marketplace"""
    # Get all komoditas for the dropdown
    komoditas = Komoditas.query.filter_by(is_deleted=False).all()

    # Check if we have initial values from harvest
    komoditas_id = request.args.get('komoditas_id')
    initial_stock = request.args.get('initial_stock', type=int)
    produksi_id = request.args.get('produksi_id', type=int)

    # Pre-select komoditas if provided
    selected_komoditas = None
    if komoditas_id:
        selected_komoditas = Komoditas.query.get(komoditas_id)

    if request.method=='POST':
        try:
            # Process product photos
            photos = request.files.getlist('product_images')
            if not photos or not photos[0].filename:
                flash('Minimal satu foto produk diperlukan', 'warning')
                return render_template('farmer/market_form.html',
                                        product=None,
                                        komoditas=komoditas,
                                        selected_komoditas=selected_komoditas,
                                        initial_stock=initial_stock)

            # Save photos and get URLs
            urls = [save_product_photo(f) for f in photos if f and f.filename]

            # Get QRIS from farmer profile
            default_qr = None
            if hasattr(current_user, 'petani_profile') and current_user.petani_profile:
                default_qr = current_user.petani_profile.qris_static

            # Create product
            p = Product(
                seller_id=current_user.id,
                nama=request.form['nama'],
                deskripsi=request.form.get('deskripsi', ''),
                komoditas_id=request.form.get('komoditas_id') or None,  # Handle empty string
                harga=Decimal(request.form['harga']),
                stok=int(request.form['stok']),
                satuan=request.form['satuan'],
                gambar_urls=",".join(urls),
                qris_static=default_qr,
                is_active=True
            )

            db.session.add(p)
            db.session.flush()  # Get the product ID without committing

            # If this product is created from a harvest, create a batch
            if produksi_id and initial_stock:
                produksi = DataPangan.query.get(produksi_id)
                if produksi and produksi.user_id == current_user.id:
                    batch = ProductBatch.create_from_harvest(
                        product_id=p.id,
                        data_pangan_id=produksi_id,
                        quantity=initial_stock,
                        notes=f"Batch awal dari panen tanggal {produksi.tanggal_panen.strftime('%d/%m/%Y')}"
                    )
                    db.session.add(batch)

            db.session.commit()

            flash('Produk berhasil ditambahkan', 'success')
            return redirect(url_for('farmer.market_manage'))

        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error adding product: {str(e)}")
            flash(f'Gagal menambahkan produk: {str(e)}', 'danger')

    return render_template('farmer/market_form.html',
                          product=None,
                          komoditas=komoditas,
                          selected_komoditas=selected_komoditas,
                          initial_stock=initial_stock)

@farmer.route('/petani/market/product/edit/<int:id>', methods=['GET','POST'])
@login_required
@role_required('petani')
def market_edit_product(id):
    """Edit an existing product"""
    # Get product
    p = Product.query.get_or_404(id)

    # Check if user owns the product
    if p.seller_id != current_user.id:
        flash('Akses ditolak', 'danger')
        return redirect(url_for('farmer.market_manage'))

    # Get all komoditas for the dropdown
    komoditas = Komoditas.query.filter_by(is_deleted=False).all()

    if request.method == 'POST':
        try:
            # Update basic product info
            p.nama = request.form['nama']
            p.deskripsi = request.form.get('deskripsi', '')
            p.komoditas_id = request.form.get('komoditas_id') or None  # Handle empty string
            p.harga = Decimal(request.form['harga'])
            p.stok = int(request.form['stok'])
            p.satuan = request.form['satuan']

            # Process new photos if uploaded
            photos = request.files.getlist('product_images')
            if photos and photos[0].filename:
                # Save new photos and get URLs
                new_urls = [save_product_photo(f) for f in photos if f and f.filename]

                # Update or append to existing photos
                if p.gambar_urls:
                    p.gambar_urls = p.gambar_urls + "," + ",".join(new_urls)
                else:
                    p.gambar_urls = ",".join(new_urls)

            # Update product
            p.updated_at = datetime.now()
            db.session.commit()

            flash('Produk berhasil diperbarui', 'success')
            return redirect(url_for('farmer.market_manage'))

        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error updating product: {str(e)}")
            flash(f'Gagal memperbarui produk: {str(e)}', 'danger')

    return render_template('farmer/market_form.html', product=p, komoditas=komoditas)

@farmer.route('/petani/market/product/delete/<int:id>', methods=['POST'])
@login_required
@role_required('petani')
def market_delete_product(id):
    p = Product.query.get_or_404(id)
    if p.seller_id == current_user.id:
        p.is_active = False
        db.session.commit()
        flash('Produk dinonaktifkan','warning')
    return redirect(url_for('farmer.market_manage'))

@farmer.route('/petani/market/product/activate/<int:id>', methods=['POST'])
@login_required
@role_required('petani')
def market_activate_product(id):
    p = Product.query.get_or_404(id)
    if p.seller_id == current_user.id:
        p.is_active = True
        db.session.commit()
        flash('Produk diaktifkan','success')
    return redirect(url_for('farmer.market_manage'))

@farmer.route('/petani/market/product/<int:id>/update-stock', methods=['POST'])
@login_required
@role_required('petani')
def market_update_product_stock(id):
    """Update product stock via AJAX"""
    try:
        product = Product.query.get_or_404(id)

        # Check if the product belongs to this seller
        if product.seller_id != current_user.id:
            return jsonify({
                'success': False,
                'message': 'Anda tidak memiliki akses ke produk ini'
            }), 403

        # Get new stock value from request
        data = request.get_json()
        new_stock = data.get('stock')

        if new_stock is None:
            return jsonify({
                'success': False,
                'message': 'Nilai stok tidak valid'
            }), 400

        try:
            new_stock = int(new_stock)
            if new_stock < 0:
                raise ValueError("Stok tidak boleh negatif")
        except ValueError:
            return jsonify({
                'success': False,
                'message': 'Nilai stok harus berupa angka positif'
            }), 400

        # Update product stock
        product.stok = new_stock
        product.updated_at = datetime.now()
        db.session.commit()

        return jsonify({
            'success': True,
            'product_id': product.id,
            'new_stock': new_stock
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating product stock: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Terjadi kesalahan: {str(e)}'
        }), 500

@farmer.route('/petani/market/order/<int:order_id>/update', methods=['POST'])
@login_required
@role_required('petani')
def market_update_order(order_id):
    """Petani bisa tandai pesanan: processed, shipped, completed."""
    order = Order.query.get_or_404(order_id)
    new_status = request.form.get('status')
    update_stock = request.form.get('update_stock') == 'yes'

    # Check if the order belongs to this seller
    order_items = OrderItem.query.filter_by(order_id=order.id).all()
    seller_ids = set(item.product.seller_id for item in order_items if item.product)

    if current_user.id not in seller_ids:
        flash('Anda tidak memiliki akses ke pesanan ini', 'danger')
        return redirect(url_for('farmer.market_manage'))

    if new_status in ('pending','paid','processed','shipped','completed','cancelled'):
        order.status = new_status

        # If order is being processed and update_stock is checked, update product stock
        if new_status == 'processed' and update_stock:
            for item in order_items:
                if item.product and item.product.seller_id == current_user.id:
                    # Get product batches ordered by harvest date (oldest first)
                    batches = ProductBatch.query.filter_by(
                        product_id=item.product.id,
                        is_deleted=False
                    ).filter(ProductBatch.quantity > 0).order_by(ProductBatch.harvest_date.asc()).all()

                    remaining_quantity = item.quantity

                    # Deduct from batches using FIFO method
                    for batch in batches:
                        if remaining_quantity <= 0:
                            break

                        if batch.quantity > 0:
                            deduct_amount = min(batch.quantity, remaining_quantity)
                            batch.quantity -= deduct_amount
                            remaining_quantity -= deduct_amount

                    # Update product stock
                    item.product.update_stock_from_batches()

                    # If no batches or not enough quantity in batches, just deduct from product stock
                    if remaining_quantity > 0:
                        item.product.stok = max(0, item.product.stok - remaining_quantity)

        db.session.commit()
        flash('Status pesanan diperbarui','success')

    # Check if this is an AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({
            'success': True,
            'order_id': order.id,
            'new_status': new_status,
            'status_badge': get_status_badge_html(new_status)
        })

    return redirect(url_for('farmer.market_manage'))

def get_status_badge_html(status):
    """Helper function to generate status badge HTML"""
    badge_class = {
        'pending': 'bg-warning',
        'paid': 'bg-success',
        'processed': 'bg-info',
        'shipped': 'bg-primary',
        'completed': 'bg-success',
        'cancelled': 'bg-danger',
        'expired': 'bg-secondary'
    }.get(status, 'bg-secondary')

    return f'<span class="badge {badge_class}">{status.capitalize()}</span>'

@farmer.route('/petani/market/order/<int:order_id>/update-ajax', methods=['POST'])
@login_required
@role_required('petani')
def market_update_order_ajax(order_id):
    """AJAX endpoint for updating order status"""
    try:
        order = Order.query.get_or_404(order_id)
        new_status = request.json.get('status')

        # Check if the order belongs to this seller
        order_items = OrderItem.query.filter_by(order_id=order.id).all()
        seller_ids = set(item.product.seller_id for item in order_items if item.product)

        if current_user.id not in seller_ids:
            return jsonify({
                'success': False,
                'message': 'Anda tidak memiliki akses ke pesanan ini'
            }), 403

        if new_status in ('pending','paid','processed','shipped','completed','cancelled'):
            order.status = new_status
            db.session.commit()

            return jsonify({
                'success': True,
                'order_id': order.id,
                'new_status': new_status,
                'status_badge': get_status_badge_html(new_status)
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Status tidak valid'
            }), 400

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating order status: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Terjadi kesalahan: {str(e)}'
        }), 500

def save_kebun_photo(photo_file) -> Tuple[bool, str]:
    """
    Save uploaded kebun photo and return status and filename
    """
    try:
        if not photo_file or photo_file.filename == '':
            return True, ''

        if not picture_allowed_file(photo_file.filename):
            return False, 'Format file tidak diizinkan!'

        filename = secrets.token_hex(8) + '_' + secure_filename(photo_file.filename)
        upload_folder = os.path.join(current_app.root_path, 'static', 'uploads', 'foto_kebun')
        os.makedirs(upload_folder, exist_ok=True)
        photo_file.save(os.path.join(upload_folder, filename))
        return True, filename

    except Exception as e:
        current_app.logger.error(f"Error saving kebun photo: {str(e)}")
        return False, 'Gagal menyimpan foto'

def validate_kebun_data(nama: str, user_id: int) -> Optional[str]:
    """
    Validate kebun data and return error message if invalid
    """
    if not nama or not nama.strip():
        return 'Nama kebun tidak boleh kosong'

    existing_kebun = Kebun.query.filter_by(nama=nama, user_id=user_id, is_deleted=False).first()
    if existing_kebun:
        return 'Nama kebun sudah ada, silakan gunakan nama lain'

    return None

@farmer.route('/api/komoditas')
def get_komoditas():
    komoditas = Komoditas.query.filter_by(is_deleted=False).all()
    return jsonify([{
        'id': k.id,
        'nama': k.nama,
        'kategori': k.kategori
    } for k in komoditas])

@farmer.route('/petani/informasi-kebun')
@login_required
@role_required('petani')
def informasi_kebun():
    """Get all active gardens for current user"""
    try:
        kebun = Kebun.query.filter_by(
            user_id=current_user.id,
            is_deleted=False
        ).order_by(Kebun.created_at.desc()).all()
        return render_template('farmer/informasi_kebun.html', kebun=kebun, page='farmer_informasi_kebun')
    except SQLAlchemyError as e:
        current_app.logger.error(f"Database error in informasi_kebun: {str(e)}")
        flash('Terjadi kesalahan saat mengambil data kebun', 'error')
        return redirect(url_for('main.dashboard'))

def handle_komoditas_input(komoditas_names):
    """
    Handle komoditas input and ensure they exist in database
    Returns list of Komoditas objects
    """
    komoditas_objects = []

    for nama in komoditas_names:
        if not nama:  # Skip empty names
            continue

        # Check if komoditas already exists
        komoditas = Komoditas.query.filter_by(
            nama=nama,
            is_deleted=False
        ).first()

        if not komoditas:
            # Create new komoditas if it doesn't exist
            komoditas = Komoditas(
                nama=nama,
                kategori='Lainnya',  # Default category, can be modified as needed
                created_by=current_user.id
            )
            db.session.add(komoditas)
            try:
                db.session.flush()  # Flush to get the ID without committing
            except SQLAlchemyError as e:
                current_app.logger.error(f"Error creating komoditas: {str(e)}")
                db.session.rollback()
                continue

        komoditas_objects.append(komoditas)

    return komoditas_objects

def create_kebun_komoditas(kebun, komoditas_list, luas_tanam=None):
    """
    Create KebunKomoditas relationships for given kebun and komoditas list
    """
    for komoditas in komoditas_list:
        kebun_komoditas = KebunKomoditas(
            kebun_id=kebun.id,
            komoditas_id=komoditas.id,
            luas_tanam=luas_tanam,  # You might want to handle individual luas_tanam later
            is_active=True,
            added_at=datetime.now()
        )
        db.session.add(kebun_komoditas)

@farmer.route('/petani/informasi-kebun/tambah-kebun', methods=['GET', 'POST'])
@login_required
@role_required('petani')
def add_kebun():
    """Add new garden with proper relationship initialization"""
    if request.method == 'POST':
        try:
            # Validate input data
            nama = request.form.get('nama_kebun')
            error_message = validate_kebun_data(nama, current_user.id)
            if error_message:
                flash(error_message, 'warning')
                return redirect(url_for('farmer.informasi_kebun'))

            # Handle komoditas input
            komoditas = request.form.getlist('komoditas[]')
            komoditas_lainnya = request.form.get('komoditas_lainnya')

            if komoditas_lainnya:
                komoditas.append(komoditas_lainnya)

            # Remove empty values and duplicates
            komoditas = list(filter(None, set(komoditas)))

            # Handle photo uploads
            uploaded_photos = request.files.getlist('fotoKebun[]')
            if not uploaded_photos or uploaded_photos[0].filename == '':
                result = 'default_thumbnail.png'
            else:
                filenames = []
                for photo in uploaded_photos:
                    if photo and photo.filename != '':
                        success, filename = save_kebun_photo(photo)
                        if not success:
                            flash(filename, 'danger')
                            return redirect(request.url)
                        filenames.append(filename)
                result = ','.join(filenames)

            # Begin transaction
            # 1. Create or get Komoditas objects
            komoditas_objects = handle_komoditas_input(komoditas)

            if not komoditas_objects:
                flash('Minimal satu komoditas harus dipilih', 'warning')
                return redirect(request.url)

            # 2. Create new Kebun
            new_kebun = Kebun(
                user_id=current_user.id,
                nama=nama,
                luas_kebun=request.form.get('luaskebun'),
                foto=result,
                komoditas=','.join(komoditas),  # Keep this for backward compatibility
                koordinat=request.form.get('koordinat'),
                unique_id=generate_unique_id()
            )
            new_kebun.users.append(current_user)

            db.session.add(new_kebun)
            db.session.flush()  # Get the ID without committing

            # 3. Create KebunKomoditas relationships
            create_kebun_komoditas(new_kebun, komoditas_objects)

            # 4. Commit all changes
            db.session.commit()

            flash('Kebun Berhasil Ditambahkan!', 'success')
            return redirect(url_for('farmer.informasi_kebun'))

        except IntegrityError as e:
            db.session.rollback()
            current_app.logger.error(f"Integrity error in add_kebun: {str(e)}")
            flash('Data kebun tidak valid', 'danger')
        except SQLAlchemyError as e:
            db.session.rollback()
            current_app.logger.error(f"Database error in add_kebun: {str(e)}")
            flash('Gagal menambahkan kebun', 'danger')
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Unexpected error in add_kebun: {str(e)}")
            flash('Terjadi kesalahan sistem', 'danger')

    return redirect(request.referrer)

@farmer.route('/petani/informasi-kebun/update/<int:id>', methods=['POST'])
@login_required
@role_required('petani')
def update_kebun(id: int):
    """Update existing garden"""
    try:
        kebun = Kebun.query.get_or_404(id)

        # Check authorization
        if current_user not in kebun.users:
            flash('Anda tidak memiliki akses untuk mengubah kebun ini', 'danger')
            return redirect(url_for('farmer.informasi_kebun'))

        # Validate new name
        new_name = request.form.get('nama_kebun')
        if kebun.nama != new_name:  # Only check if name is being changed
            error_message = validate_kebun_data(new_name, current_user.id)
            if error_message:
                flash(error_message, 'warning')
                return redirect(url_for('farmer.informasi_kebun'))
        kebun.nama = new_name

        # Update luas kebun
        kebun.luas_kebun = request.form.get('luaskebun')

        # Update koordinat
        kebun.koordinat = request.form.get('koordinat')

        # Update komoditas
        komoditas = request.form.getlist('komoditas[]')
        komoditas_lainnya = request.form.get('komoditas_lainnya')
        if komoditas_lainnya:
            komoditas.append(komoditas_lainnya)
        kebun.komoditas = ','.join(komoditas)

        # Optionally, handle photo uploads if you want to allow updating photos
        # For simplicity, we can skip updating photos here

        db.session.commit()
        flash('Data Kebun Berhasil diubah!', 'success')
        return redirect(url_for('farmer.informasi_kebun'))

    except IntegrityError:
        db.session.rollback()
        flash('Data kebun tidak valid', 'danger')
    except SQLAlchemyError as e:
        db.session.rollback()
        current_app.logger.error(f"Database error in update_kebun: {str(e)}")
        flash('Gagal mengubah data kebun', 'danger')
    return redirect(url_for('farmer.informasi_kebun'))

@farmer.route('/petani/informasi-kebun/import_kebun', methods=['POST'])
@login_required
@role_required('petani')
def import_kebun():
    """Import multiple gardens from Excel file"""
    if 'excel_file' not in request.files:
        flash('Tidak ada file yang dipilih!', 'error')
        return redirect(request.url)

    excel_file = request.files['excel_file']
    if excel_file.filename == '' or not report_allowed_file(excel_file.filename):
        flash('File tidak valid. Unggah file Excel (.xlsx)!', 'error')
        return redirect(request.url)

    try:
        wb = load_workbook(excel_file)
        sheet = wb.active

        # Begin transaction
        added_kebun = 0
        skipped_kebun = 0

        for row in sheet.iter_rows(min_row=2):
            nama_kebun = row[0].value
            if not nama_kebun:  # Skip empty rows
                continue

            # Validate garden name
            error_message = validate_kebun_data(nama_kebun, current_user.id)
            if error_message:
                skipped_kebun += 1
                continue

            # Create new garden
            new_kebun = Kebun(
                nama=nama_kebun,
                koordinat=f"{row[2].value}, {row[1].value}",  # longitude, latitude
                luas_kebun=row[3].value,
                user_id=current_user.id,
                unique_id=generate_unique_id()
            )
            new_kebun.users.append(current_user)
            db.session.add(new_kebun)
            added_kebun += 1

        db.session.commit()

        message = f'Berhasil mengimpor {added_kebun} kebun'
        if skipped_kebun > 0:
            message += f' ({skipped_kebun} kebun dilewati karena duplikat)'
        flash(message, 'success')

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error importing kebun: {str(e)}")
        flash('Gagal mengimpor data. Pastikan format file sesuai template.', 'error')

    return redirect(url_for('farmer.informasi_kebun'))

@farmer.route('/petani/informasi-kebun/delete_kebun/<int:id>', methods=['POST', 'GET'])
@login_required
@role_required('petani')
def del_kebun(id):
    """Soft delete garden"""
    try:
        kebun = Kebun.query.get_or_404(id)

        # Mengecek authorization
        if current_user not in kebun.users:
            flash('Anda tidak diizinkan menghapus kebun ini', 'danger')
            return redirect(url_for('farmer.informasi_kebun'))

        # Soft 'delete' untuk data kebun
        kebun.is_deleted = True
        kebun.deleted_at = datetime.now()

        # Menghapus semua user yang terkait
        kebun.users.clear()

        db.session.commit()
        flash('Kebun Berhasil Dihapus!', 'danger')

    except SQLAlchemyError as e:
        db.session.rollback()
        current_app.logger.error(f"Database error in del_kebun: {str(e)}")
        flash('Gagal menghapus kebun ', 'danger')

    return redirect(url_for('farmer.informasi_kebun'))

@farmer.route('/petani/produksi/tambah', methods=['GET', 'POST'])
@login_required
@role_required('petani')
def tambah_produksi():
    if request.method == 'POST':
        try:
            kebun_id = request.form.get('kebun_id')
            komoditas_id = request.form.get('komoditas_id')
            jml_bibit = request.form.get('jml_bibit')
            tanggal_bibit = request.form.get('tanggal_bibit')

            # Validasi input
            if not kebun_id or not komoditas_id or not jml_bibit or not tanggal_bibit:
                flash('Semua kolom harus diisi!', 'warning')
                return redirect(request.url)

            # Konversi dan validasi tanggal
            try:
                tanggal_bibit = datetime.strptime(tanggal_bibit, '%Y-%m-%d')
                # Hitung estimasi panen 60 hari setelah tanggal bibit
                estimasi_panen = tanggal_bibit + timedelta(days=60)
            except ValueError:
                flash('Format tanggal tidak valid!', 'warning')
                return redirect(request.url)

            # Validasi jumlah bibit
            if not jml_bibit.isdigit() or int(jml_bibit) <= 0:
                flash('Jumlah bibit harus berupa angka positif!', 'warning')
                return redirect(request.url)

            produksi = DataPangan(
                kebun_id=kebun_id,
                komoditas_id=komoditas_id,
                user_id=current_user.id,
                jml_bibit=int(jml_bibit),
                tanggal_bibit=tanggal_bibit,
                estimasi_panen=estimasi_panen,
                status='Penanaman'
            )
            db.session.add(produksi)
            db.session.commit()
            flash('Data produksi berhasil ditambahkan!', 'success')
            return redirect(url_for('farmer.manajemen_produksi'))
        except SQLAlchemyError as e:
            db.session.rollback()
            current_app.logger.error(f"Database error in tambah_produksi: {str(e)}")
            flash('Gagal menambahkan data produksi', 'danger')
    return redirect(request.referrer)

@farmer.route('/petani/produksi/update_produksi/<int:id>', methods=['POST', 'GET'])
@login_required
@role_required('petani')
def update_produksi(id):
    produksi = DataPangan.query.get_or_404(id)
    if request.method == 'POST':
        try:
            # Ambil nilai dari form
            jml_bibit = request.form.get('update_jml_bibit')
            tanggal_bibit = request.form.get('update_tanggal_bibit')
            estimasi_panen = request.form.get('update_estimasi_panen')

            # Konversi tanggal bibit
            produksi.tanggal_bibit = datetime.strptime(tanggal_bibit, '%Y-%m-%d')

            # Validasi dan konversi estimasi panen jika ada
            if estimasi_panen:
                produksi.estimasi_panen = datetime.strptime(estimasi_panen, '%Y-%m-%d')

            # Update status dan jumlah bibit
            produksi.jml_bibit = jml_bibit

            db.session.commit()
            flash('Data produksi berhasil diperbarui!', 'success')
            return redirect(url_for('farmer.manajemen_produksi'))  # Kembali ke halaman manajemen produksi
        except SQLAlchemyError as e:
            db.session.rollback()
            current_app.logger.error(f"Database error in update_produksi: {str(e)}")
            flash('Gagal memperbarui data produksi', 'danger')
    return redirect(url_for('farmer.manajemen_produksi'))  # Jika bukan POST, kembali ke halaman manajemen produksi

@farmer.route('/petani/produksi/update_panen/<int:id>', methods=['POST', 'GET'])
@login_required
@role_required('petani')
def update_panen(id):
    produksi = DataPangan.query.get_or_404(id)
    if request.method == 'POST':
        try:
            jml_panen = request.form.get('hasil_panen')
            tanggal_panen = request.form.get('tanggal_panen')
            update_stock = request.form.get('update_stock') == 'yes'

            # Mengubah format tanggal panen
            produksi.tanggal_panen = datetime.strptime(tanggal_panen, '%Y-%m-%d')  # Pastikan format tanggal sesuai

            # Update status dan jumlah panen
            produksi.jml_panen = jml_panen
            produksi.status = 'Panen'

            db.session.commit()
            flash('Data produksi berhasil diperbarui!', 'success')

            # Jika user ingin update stok produk
            if update_stock:
                # Cari produk dengan komoditas yang sama
                matching_products = Product.query.filter_by(
                    seller_id=current_user.id,
                    komoditas_id=produksi.komoditas_id,
                    is_active=True
                ).all()

                if matching_products:
                    # Jika ada produk yang cocok, tampilkan halaman untuk memilih produk
                    return render_template(
                        'farmer/update_product_stock.html',
                        products=matching_products,
                        harvest_amount=jml_panen,
                        produksi_id=produksi.id,
                        produksi=produksi
                    )
                else:
                    # Jika tidak ada produk yang cocok, tawarkan untuk membuat produk baru
                    flash('Panen berhasil dicatat! Anda belum memiliki produk untuk komoditas ini. Ingin membuat produk baru?', 'info')
                    return redirect(url_for('farmer.market_add_product', komoditas_id=produksi.komoditas_id, initial_stock=jml_panen))

            return redirect(url_for('farmer.manajemen_produksi'))  # Kembali ke halaman manajemen produksi
        except SQLAlchemyError as e:
            db.session.rollback()
            current_app.logger.error(f"Database error in update_panen: {str(e)}")
            flash('Gagal memperbarui data produksi', 'danger')
    return redirect(url_for('farmer.manajemen_produksi'))  # Jika bukan POST, kembali ke halaman manajemen produksi


@farmer.route('/petani/produksi/delete', methods=['POST', 'GET'])
@login_required
@role_required('petani')
def delete_produksi():
    try:
        delete_ids = request.form.getlist('check')  # Ambil daftar ID dari checkbox yang dipilih
        if not delete_ids:
            flash('Tidak ada data yang dipilih untuk dihapus', 'warning')
            return redirect(url_for('farmer.manajemen_produksi'))

        # Soft delete untuk semua ID yang dipilih
        DataPangan.query.filter(
            DataPangan.id.in_(delete_ids),
            DataPangan.user_id == current_user.id  # Pastikan hanya data milik user yang bisa dihapus
        ).update({DataPangan.is_deleted: True}, synchronize_session=False)

        db.session.commit()
        flash('Data produksi berhasil dihapus!', 'success')
    except SQLAlchemyError as e:
        db.session.rollback()
        current_app.logger.error(f"Database error in delete_produksi: {str(e)}")
        flash('Gagal menghapus data produksi', 'danger')
    return redirect(url_for('farmer.manajemen_produksi'))

@farmer.route('/petani/produksi/delete/<int:id>', methods=['POST', 'GET'])
@login_required
@role_required('petani')
def delete_single_produksi(id):
    try:
        produksi = DataPangan.query.get_or_404(id)  # Pastikan ID valid
        produksi.is_deleted = True  # Soft delete
        # db.session.delete(produksi)  # Hapus baris ini jika menggunakan soft delete
        db.session.commit()
        flash('Data produksi berhasil dihapus!', 'warning')
    except SQLAlchemyError as e:
        db.session.rollback()
        current_app.logger.error(f"Database error in delete_produksi: {str(e)}")
        flash('Gagal menghapus data produksi', 'danger')
    return redirect(url_for('farmer.manajemen_produksi'))

@farmer.route('/petani/produksi/import-data', methods=['POST'])
@login_required
@role_required('petani')
def import_produksi():
    """Import production data from Excel file or CSV file"""
    if 'excel_file' not in request.files:
        flash('Tidak ada file yang dipilih!', 'error')
        return redirect(request.url)

    file = request.files['excel_file']
    import_type = request.form.get('import_type', 'penanaman')

    if file.filename == '':
        flash('Tidak ada file yang dipilih!', 'error')
        return redirect(request.url)

    # Validate file size
    if len(file.read()) > MAX_CONTENT_LENGTH:
        flash('Ukuran file terlalu besar! Maksimal 5MB', 'error')
        return redirect(request.url)
    file.seek(0)  # Reset file pointer

    # Check file extension
    file_extension = os.path.splitext(file.filename)[1].lower()
    is_csv = file_extension == '.csv'
    is_excel = file_extension in ['.xlsx', '.xls']
    
    if not (is_csv or is_excel):
        flash('File tidak valid. Unggah file Excel (.xlsx, .xls) atau CSV (.csv)!', 'error')
        return redirect(request.url)

    # Validate filename matches import type - make this optional
    file_name_without_extension = os.path.splitext(file.filename)[0].lower()
    if import_type not in file_name_without_extension:
        flash(f'Peringatan: Sebaiknya nama file mengandung kata "{import_type}" untuk pemahaman lebih jelas.', 'warning')
        # Continue anyway, don't redirect

    try:
        # Process file based on type
        if is_csv:
            # Handle CSV file
            df = pd.read_csv(file, encoding='utf-8')
        else:
            # Handle Excel file
            df = pd.read_excel(file)
        
        # Lowercase and standardize column names
        df.columns = [col.lower().strip() for col in df.columns]
        
        # Begin transaction
        added_records = 0
        skipped_records = 0

        for index, row in df.iterrows():
            try:
                # Basic validation for required fields
                if 'kebun' not in df.columns or 'komoditas' not in df.columns or 'jml_bibit' not in df.columns:
                    flash('Format file tidak sesuai dengan template. Pastikan memiliki kolom kebun, komoditas, dan jml_bibit.', 'error')
                    return redirect(request.url)
                
                kebun_id = row.get('kebun')
                komoditas_name = row.get('komoditas')
                jml_bibit = row.get('jml_bibit')
                
                # Parse planting date with multiple formats
                tanggal_bibit = None
                tanggal_bibit_value = row.get('tanggal_penanaman')
                
                # Try multiple date formats
                if tanggal_bibit_value:
                    for date_format in ['%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m-%d-%Y']:
                        try:
                            if isinstance(tanggal_bibit_value, str):
                                tanggal_bibit = datetime.strptime(tanggal_bibit_value, date_format)
                                break
                            elif isinstance(tanggal_bibit_value, datetime):
                                tanggal_bibit = tanggal_bibit_value
                                break
                        except ValueError:
                            continue

                if not all([kebun_id, komoditas_name, jml_bibit, tanggal_bibit]):
                    current_app.logger.warning(f"Skipping row due to missing required fields: {row}")
                    skipped_records += 1
                    continue

                # Find kebun
                kebun = Kebun.query.filter_by(
                    id=int(kebun_id) if not isinstance(kebun_id, int) else kebun_id,
                    user_id=current_user.id,
                    is_deleted=False
                ).first()

                if not kebun:
                    current_app.logger.warning(f"Kebun tidak ditemukan: {kebun_id}")
                    skipped_records += 1
                    continue

                # Find komoditas (try by name/string instead of just ID)
                komoditas = None
                if isinstance(komoditas_name, int) or komoditas_name.isdigit():
                    komoditas = Komoditas.query.filter_by(
                        id=int(komoditas_name),
                        is_deleted=False
                    ).first()
                else:
                    komoditas = Komoditas.query.filter(
                        Komoditas.nama.ilike(f"%{komoditas_name}%"),
                        Komoditas.is_deleted==False
                    ).first()

                if not komoditas:
                    current_app.logger.warning(f"Komoditas tidak ditemukan: {komoditas_name}")
                    skipped_records += 1
                    continue

                # Create base production record
                produksi = DataPangan(
                    kebun_id=kebun.id,
                    komoditas_id=komoditas.id,
                    user_id=current_user.id,
                    jml_bibit=int(jml_bibit) if not isinstance(jml_bibit, int) else jml_bibit,
                    tanggal_bibit=tanggal_bibit,
                    estimasi_panen=tanggal_bibit + timedelta(days=60)
                )

                # Handle panen data if applicable
                if import_type == 'panen':
                    jml_panen = row.get('jml_panen')
                    tanggal_panen_value = row.get('tanggal_panen')
                    
                    # Try multiple date formats for harvest date
                    tanggal_panen = None
                    if tanggal_panen_value:
                        for date_format in ['%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m-%d-%Y']:
                            try:
                                if isinstance(tanggal_panen_value, str):
                                    tanggal_panen = datetime.strptime(tanggal_panen_value, date_format)
                                    break
                                elif isinstance(tanggal_panen_value, datetime):
                                    tanggal_panen = tanggal_panen_value
                                    break
                            except ValueError:
                                continue

                    if not jml_panen or not tanggal_panen:
                        current_app.logger.warning(f"Missing harvest data for panen record: {row}")
                        skipped_records += 1
                        continue

                    produksi.jml_panen = int(jml_panen) if not isinstance(jml_panen, int) else jml_panen
                    produksi.tanggal_panen = tanggal_panen
                    produksi.status = 'Panen'
                else:
                    produksi.status = 'Penanaman'

                db.session.add(produksi)
                added_records += 1

            except Exception as row_error:
                current_app.logger.error(f"Error processing row: {str(row_error)}")
                skipped_records += 1
                continue

        db.session.commit()

        message = f'Berhasil mengimpor {added_records} data produksi'
        if skipped_records > 0:
            message += f' ({skipped_records} data dilewati karena tidak valid)'
        flash(message, 'success')

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error importing production data: {str(e)}")
        flash('Gagal mengimpor data. Pastikan format file sesuai template.', 'error')

    return redirect(url_for('farmer.manajemen_produksi'))

@farmer.route('/api/download-template/<template_type>')
@login_required
@role_required('petani')
def download_template(template_type):
    """Download Excel template for data import"""
    try:
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # Add headers based on template type
        if template_type == 'penanaman':
            headers = ['ID Kebun', 'ID Komoditas', 'Jumlah Bibit', 'Tanggal Bibit (YYYY-MM-DD)']
        elif template_type == 'panen':
            headers = ['ID Kebun', 'ID Komoditas', 'Jumlah Bibit', 'Tanggal Bibit (YYYY-MM-DD)',
                       'Jumlah Panen', 'Tanggal Panen (YYYY-MM-DD)']
        else:
            return jsonify({'error': 'Template type not valid'}), 400

        # Write headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Create helper sheet
        help_sheet = wb.create_sheet(title="Bantuan")

        # Get user's active gardens
        kebun_list = Kebun.query.filter_by(
            user_id=current_user.id,
            is_deleted=False
        ).all()

        # Get active commodities
        komoditas_list = Komoditas.query.filter_by(
            is_deleted=False
        ).all()

        # Write garden list
        help_sheet['A1'] = 'Daftar Kebun:'
        help_sheet['A2'] = 'ID Kebun'
        help_sheet['B2'] = 'Nama Kebun'

        for i, kebun in enumerate(kebun_list, 3):
            help_sheet[f'A{i}'] = kebun.id
            help_sheet[f'B{i}'] = kebun.nama

        # Write commodity list
        row_start = len(kebun_list) + 5
        help_sheet[f'A{row_start}'] = 'Daftar Komoditas:'
        help_sheet[f'A{row_start+1}'] = 'ID Komoditas'
        help_sheet[f'B{row_start+1}'] = 'Nama Komoditas'

        for i, komoditas in enumerate(komoditas_list, row_start+2):
            help_sheet[f'A{i}'] = komoditas.id
            help_sheet[f'B{i}'] = komoditas.nama

        # Add instructions
        row_start = row_start + len(komoditas_list) + 3
        help_sheet[f'A{row_start}'] = 'Petunjuk Pengisian:'
        instructions = [
            '1. Tanggal dalam format YYYY-MM-DD (contoh: 2025-01-19)',
            '2. Jumlah bibit dan panen harus angka positif',
            '3. ID Kebun harus sesuai dengan daftar kebun di atas',
            '4. ID Komoditas harus sesuai dengan daftar komoditas di atas',
            '5. Field wajib tidak boleh kosong',
            f'6. Status akan otomatis diisi {"Panen" if template_type == "panen" else "Penanaman"}'
        ]

        for i, instruction in enumerate(instructions, row_start+1):
            help_sheet[f'A{i}'] = instruction

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Gunakan parameter yang kompatibel dengan versi Flask yang lebih lama
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            attachment_filename=f'template_{template_type}.xlsx'  # Gunakan attachment_filename, bukan download_name
        )

    except Exception as e:
        # Tambahkan logging yang lebih detail
        current_app.logger.error(f"Error generating template: {str(e)}")
        current_app.logger.error(f"Error traceback: {traceback.format_exc()}")
        return jsonify({'error': 'Failed to generate template'}), 500

@farmer.route('/api/produksi', methods=['GET'])
@login_required
@role_required('petani')
def api_produksi():
    field_id = request.args.get('field')
    commodity_id = request.args.get('commodity')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    three_months_ago = datetime.now() - timedelta(days=90)
    query = DataPangan.query.filter(
        DataPangan.user_id == current_user.id,
        DataPangan.is_deleted == False,
        DataPangan.tanggal_bibit >= three_months_ago
    )

    # Apply filters
    if field_id:
        query = query.filter(DataPangan.kebun_id == field_id)
    if commodity_id:
        query = query.filter(DataPangan.komoditas_id == commodity_id)
    if start_date:
        query = query.filter(DataPangan.tanggal_bibit >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(DataPangan.tanggal_bibit <= datetime.strptime(end_date, '%Y-%m-%d'))

    produksi = query.all()

    chart_data = []
    for prod in produksi:
        if prod.tanggal_panen and prod.jml_panen:
            chart_data.append({
                'kebun': prod.kebun.nama,
                'komoditas': prod.komoditas_info.nama,
                'tanggal_panen': prod.tanggal_panen.strftime('%Y-%m-%d'),
                'hasil_panen': prod.jml_panen
            })

    return jsonify(chart_data)

@farmer.route('/api/produksi', methods=['GET', 'POST'])
@login_required
@role_required('petani')
def get_produksi_data():
    # Get filter parameters
    field_id = request.args.get('field')
    commodity_id = request.args.get('commodity')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    # Base query
    query = DataPangan.query.filter(
        DataPangan.user_id == current_user.id,
        DataPangan.is_deleted == False
    )

    # Apply filters
    if field_id:
        query = query.filter(DataPangan.kebun_id == field_id)
    if commodity_id:
        query = query.filter(DataPangan.komoditas_id == commodity_id)
    if start_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            query = query.filter(DataPangan.tanggal_bibit >= start)
        except ValueError:
            pass
    if end_date:
        try:
            end = datetime.strptime(end_date, '%Y-%m-%d')
            query = query.filter(DataPangan.tanggal_bibit <= end)
        except ValueError:
            pass

    # Get data
    produksi = query.all()

    # Format data
    data = []
    for p in produksi:
        data.append({
            'id': p.id,
            'kebun': p.kebun.nama,
            'komoditas': p.komoditas_info.nama,
            'tanggal_bibit': p.tanggal_bibit.strftime('%Y-%m-%d') if p.tanggal_bibit else None,
            'jml_bibit': p.jml_bibit,
            'tanggal_panen': p.tanggal_panen.strftime('%Y-%m-%d') if p.tanggal_panen else None,
            'jml_panen': p.jml_panen,
            'status': p.status
        })

    return jsonify(data)

@farmer.route('/petani/produksi/cetak-laporan', methods=['GET'])
@login_required
@role_required('petani')
def cetak_laporan_produksi():
    try:
        # Query data produksi
        three_months_ago = datetime.now() - timedelta(days=90)
        produksi = DataPangan.query.filter(
            DataPangan.user_id == current_user.id,
            DataPangan.is_deleted == False,
            DataPangan.tanggal_bibit >= three_months_ago
        ).order_by(DataPangan.tanggal_bibit.desc()).all()

        # Buat buffer untuk PDF
        buffer = BytesIO()

        # Buat dokumen PDF
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )

        # Daftar elemen yang akan ditambahkan ke PDF
        elements = []

        # Tambahkan judul
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30
        )
        elements.append(Paragraph("Laporan Hasil Penanaman - Panen", title_style))

        # Persiapkan data untuk tabel
        data = [['Kebun', 'Jumlah Bibit', 'Tanggal Bibit', 'Status', 'Hasil Panen']]

        for p in produksi:
            hasil_panen = f"{p.jml_panen}kg, {format_date(p.tanggal_panen, format='d MMM Y', locale='id')}" if p.jml_panen and p.tanggal_panen else "Belum panen"
            data.append([
                p.kebun.nama,
                str(p.jml_bibit),
                format_date(p.tanggal_bibit, format='d MMM Y', locale='id') if p.tanggal_bibit else "-",
                p.status,
                hasil_panen
            ])

        # Buat tabel
        table = Table(data, colWidths=[120, 80, 100, 80, 120])

        # Style tabel
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))

        elements.append(table)

        # Buat PDF
        doc.build(elements)

        # Siapkan file untuk didownload
        buffer.seek(0)
        return send_file(
            buffer,
            download_name=f'laporan_produksi_{datetime.now().strftime("%Y%m%d")}.pdf',
            mimetype='application/pdf'
        )

    except Exception as e:
        current_app.logger.error(f"Error generating PDF: {str(e)}")
        flash('Gagal membuat laporan PDF', 'error')
        return redirect(url_for('farmer.manajemen_produksi'))

@farmer.route('/petani/produksi/update-product-stock', methods=['POST'])
@login_required
@role_required('petani')
def update_product_stock_from_harvest():
    """Update product stock from harvest data"""
    if request.method == 'POST':
        try:
            produksi_id = request.form.get('produksi_id')
            product_id = request.form.get('product_id')
            quantity = request.form.get('quantity', type=int)
            notes = request.form.get('notes')

            # Validate inputs
            if not produksi_id or not product_id or not quantity:
                flash('Semua kolom wajib diisi', 'warning')
                return redirect(url_for('farmer.manajemen_produksi'))

            # Get the harvest and product data
            produksi = DataPangan.query.get_or_404(produksi_id)
            product = Product.query.get_or_404(product_id)

            # Verify ownership
            if produksi.user_id != current_user.id or product.seller_id != current_user.id:
                flash('Anda tidak memiliki akses ke data ini', 'danger')
                return redirect(url_for('farmer.manajemen_produksi'))

            # Create a new batch
            batch = ProductBatch.create_from_harvest(
                product_id=product.id,
                data_pangan_id=produksi.id,
                quantity=quantity,
                notes=notes
            )

            db.session.add(batch)

            # Update product stock
            product.stok += quantity

            db.session.commit()

            flash(f'Stok produk {product.nama} berhasil diperbarui dengan {quantity} {product.satuan} dari hasil panen', 'success')
            return redirect(url_for('farmer.market_manage'))

        except ValueError as e:
            flash(f'Error: {str(e)}', 'danger')
            return redirect(url_for('farmer.manajemen_produksi'))
        except SQLAlchemyError as e:
            db.session.rollback()
            current_app.logger.error(f"Database error in update_product_stock_from_harvest: {str(e)}")
            flash('Gagal memperbarui stok produk', 'danger')
            return redirect(url_for('farmer.manajemen_produksi'))

    return redirect(url_for('farmer.manajemen_produksi'))

@farmer.route('/api/export-csv', methods=['GET'])
@login_required
@role_required('petani')
def export_csv():
    # Similar query as api_produksi but returns CSV formatted data
    query = DataPangan.query.filter(
        DataPangan.user_id == current_user.id,
        DataPangan.is_deleted == False
    )

    produksi = query.all()

    csv_data = []
    for prod in produksi:
        csv_data.append({
            'kebun': prod.kebun.nama,
            'komoditas': prod.komoditas_info.nama,
            'tanggal_bibit': prod.tanggal_bibit.strftime('%Y-%m-%d') if prod.tanggal_bibit else '',
            'jml_bibit': prod.jml_bibit,
            'tanggal_panen': prod.tanggal_panen.strftime('%Y-%m-%d') if prod.tanggal_panen else '',
            'jml_panen': prod.jml_panen if prod.jml_panen else ''
        })

    return jsonify(csv_data)
